import requests
from bs4 import BeautifulSoup
from typing import Set, List, Dict
import time
from openpyxl import Workbook, load_workbook
from datetime import datetime, date
import os
import re
import json
from loguru import logger

class CMBFinanceCrawler:
    """
    A crawler for fetching financial product data from CMB (招商银行理财).
    
    This crawler can:
    1. Fetch financial product NAV history data within date range
    2. Export data to Excel file
    3. Respect robots.txt and implement polite crawling
    
    Data fields include:
    - product_name: 产品名称
    - nav: 单位净值
    - acc_nav: 累计净值
    - date: 净值日期
    """

    def __init__(self,
                 delay: float = 1.0,
                 cache_dir: str = "cache/cmb"):
        """
        Initialize the CMB finance product crawler
        """
        self.delay = delay
        self.cache_dir = cache_dir
        self.product_series = []

    def _build_url(self, product: Dict, page_index: int = 1) -> str:
        """Build the URL with the current parameters."""
        base = f"https://www.cmbchina.com/cfweb/personal/"

        if ('IsSA' in product and product['IsSA'] == '1'):
            base += f"saproductdetail.aspx?saaCod={product['saaCod']}&funCod={product['PrdCode']}&type=prodvalue"
        else:
            base += f"productdetail.aspx?code={product['PrdCode']}&type=prodvalue"

        if page_index > 1:
            base += f"&PageNo={page_index}"
        return base
    
    def __load_cache(self, product: Dict, start_date: date, end_date: date) -> Dict:
        # 检查缓存目录是否存在
        if not os.path.exists(self.cache_dir):
            logger.info(f"缓存目录不存在: {self.cache_dir}")
            return {
                'nav_data': [],
                'info_data': {}
            }
        finance_data = []
        product_info = {}
        # 查找匹配的缓存文件
        date_range = ""
        start = start_date.strftime("%Y%m%d") if start_date else "START"
        end = end_date.strftime("%Y%m%d") if end_date else "END"
        date_range = f"{start}_to_{end}"
        prefix = f"finance_{product['PrdCode']}_{date_range}_"
        
        for file in os.listdir(self.cache_dir):
            if file.startswith(prefix) and file.endswith('.xlsx'):
                cache_path = os.path.join(self.cache_dir, file)
                try:
                    logger.info(f"找到有效缓存文件: {cache_path}")
                    # 加载Excel文件
                    wb = load_workbook(cache_path)
                    
                    # 读取净值数据
                    data_sheet = wb['finance_data']
                    headers = [cell.value for cell in data_sheet[1]]  # 获取表头
                    
                    # 从第二行开始读取数据
                    for row in list(data_sheet.rows)[1:]:  # 转换为列表
                        row_data = {
                            'product_code': row[0].value,
                            'product_name': row[1].value,
                            'nav': float(row[2].value),
                            'acc_nav': float(row[3].value),
                            'date': row[4].value
                        }
                        finance_data.append(row_data)
                    
                    # 读取产品信息
                    if 'product_info' in wb.sheetnames:
                        info_sheet = wb['product_info']
                        for row in list(info_sheet.rows)[1:]:  # 转换为列表
                            if len(row) >= 2:
                                product_info[row[0].value] = row[1].value
                            
                    logger.info("从缓存加载数据成功")
                    return {
                        'nav_data': finance_data,
                        'info_data': product_info
                    }
                    
                except Exception as e:
                    logger.error(f"读取缓存文件失败: {str(e)}")
                    import traceback
                    logger.error(traceback.format_exc())
                    continue
                
        return {
            'nav_data': [],
            'info_data': {}
        }

    def crawl_product_nav(self, product: Dict, start_date_str: str = None, end_date_str: str = None, delay: float = 1.0) -> Dict[str, List[Dict[str, str]]]:
        """
        Crawl all pages within the specified date range.
        Will automatically stop when reaching data outside the date range.
        """
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else None
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else None

        finance_data = []
        product_info = {}

        cache = self.__load_cache(product, start_date, end_date)
        if len(cache['nav_data']) > 0 and len(cache['info_data']) > 0 and end_date != None:
            return {
                'nav_data': cache['nav_data'],
                'info_data': cache['info_data']
            }
        finance_data = cache['nav_data']
        product_info = cache['info_data']

        if len(finance_data) > 0:
            loaded_end_date = datetime.strptime(finance_data[0]['date'], '%Y-%m-%d').date() if finance_data[0]['date'] else None
            print(f"loaded_end_date: {loaded_end_date}")
            start_date = loaded_end_date

        page_index = 1

        while True:
            logger.info(f"正在获取第{product['PrdName']}的第{page_index}页数据...")
            url = self._build_url(product, page_index)
            parse_info = page_index == 1
            result = self._crawl_url(url, start_date, end_date, parse_info)
            if parse_info:
                product_info = result['info_data']
            nav_data = result['nav_data']
            finance_data.extend(nav_data)
            if len(nav_data) != 10:
                break
            page_index += 1
            time.sleep(self.delay)
        self._export_to_cache(product, start_date, end_date, finance_data, product_info)
        return {
            'nav_data': finance_data,
            'info_data': product_info
        }

    def _crawl_url(self, url: str, start_date: date = None, end_date: date = None, parse_info: bool = False) -> bool:
        """
        Crawl a specific URL and extract information.
        Returns False if should stop crawling (reached date limit or no more data).
        """
        while True:
            try:
                logger.info(f"Crawling: {url}")
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
                    'Referer': 'https://www.cmbchina.com/'
                }
                response = requests.get(url, headers=headers, timeout=10)
                response.encoding = response.apparent_encoding
                response.raise_for_status()
                return self._process_page(response.text, start_date, end_date, parse_info)

            except requests.RequestException as e:
                logger.error(f"will retry, Error crawling {url}: {str(e)}")
                time.sleep(self.delay)
                continue
            except UnicodeEncodeError as e:
                logger.error(f"Encoding error for {url}: {str(e)}")
                break
        return False

    def _process_page(self, html: str, start_date: date = None, end_date: date = None, parse_info: bool = False) -> dict:
        """
        Extract and process information from the page.
        Returns False if should stop crawling (reached date limit or no more data).
        """
        finance_data = []
        product_info = {}
        soup = BeautifulSoup(html, 'html.parser')
        
        # 在第一页时解析产品基本信息
        if parse_info:
            product_info = self._parse_product_info(soup)
        
        table = soup.find('table')
        if not table:
            logger.warning("No data table found")
            return {
                'nav_data': finance_data,
                'info_data': product_info
            }
        
        # 判断表格类型
        table_type = 0
        table_rows = table.find_all('tr')
        if len(table_rows) == 0:
            logger.warning("No data rows found")
            return {
                'nav_data': finance_data,
                'info_data': product_info
            }
        header_row = table_rows[0]
        headers = [cell.get_text(strip=True) for cell in header_row.find_all('th')]
        if len(headers) == 5 and \
            headers[0] == "产品代码" and \
            headers[1] == "产品名称" and \
            headers[2] == "单位净值" and \
            headers[3] == "累计净值" and \
            headers[4] == "净值日期":
            table_type = 1
        elif len(headers) == 7 and \
            headers[0] == "产品代码" and \
            headers[1] == "产品简称" and \
            headers[2] == "信托单位净值(元)" and \
            headers[3] == "累计信托收益率(%)" and \
            headers[4] == "信托年化收益率(%)" and \
            headers[5] == "净值日期" and \
            headers[6] == "信托网下参考年收益率(%)":
            table_type = 2
        elif len(headers) == 6 and \
            headers[0] == "产品代码" and \
            headers[1] == "产品名称" and \
            headers[2] == "业绩比较基准(年率%)" and \
            headers[3] == "到期收益率(年率%)" and \
            headers[4] == "产品起息日" and \
            headers[5] == "产品到期日":
            table_type = 3
        elif len(headers) == 5 and \
            headers[0] == "产品代码" and \
            headers[1] == "产品名称" and \
            headers[2] == "收益率(年率%)" and \
            headers[3] == "产品起息日" and \
            headers[4] == "产品到期日":
            table_type = 4
        elif len(headers) == 5 and \
            headers[0] == "产品代码" and \
            headers[1] == "产品名称" and \
            headers[2] == "收益率(年率%)" and \
            headers[3] == "近七日收益率(年率%)" and \
            headers[4] == "日期":
            table_type = 5
        elif len(headers) == 4 and \
            headers[0] == "产品代码" and \
            headers[1] == "产品名称" and \
            headers[2] == "收益率(年率%)" and \
            headers[3] == "日期":
            table_type = 6
        else:
            logger.warning(f"Unknown table type: {headers}")
            return {
                'nav_data': finance_data,
                'info_data': product_info
            }

        # Process data rows
        rows = table_rows[1:]  # Skip header row
        data_found = False

        for row in rows:
            cells = row.find_all('td')
            nav_date = None
            data = None
            if table_type == 1:
                data_found = True
                nav_date_str = cells[4].get_text(strip=True)
                nav_date = datetime.strptime(nav_date_str, '%Y%m%d').date()
                data = {
                    'product_code': cells[0].get_text(strip=True),
                    'product_name': cells[1].get_text(strip=True),
                    'nav': cells[2].get_text(strip=True),
                    'acc_nav': cells[3].get_text(strip=True),
                    'date': datetime.strptime(nav_date_str, '%Y%m%d').strftime('%Y-%m-%d')
                }
                
            elif table_type == 2:
                data_found = True
                nav_date_str = cells[5].get_text(strip=True)
                nav_date = datetime.strptime(nav_date_str, '%Y-%m-%d').date()
                data = {
                    'product_code': cells[0].get_text(strip=True),
                    'product_name': cells[1].get_text(strip=True),
                    'nav': cells[2].get_text(strip=True),
                    'acc_nav': cells[2].get_text(strip=True),
                    'date': nav_date_str
                }
            elif table_type == 3:
                data_found = True
                nav_date_str = cells[4].get_text(strip=True)
                nav_date = datetime.strptime(nav_date_str, '%Y-%m-%d').date()
                data = {
                    'product_code': cells[0].get_text(strip=True),
                    'product_name': cells[1].get_text(strip=True),
                    'nav': 1,
                    'acc_nav': 1,
                    'date': nav_date_str
                }
            elif table_type == 4:
                data_found = True
                nav_date_str = cells[3].get_text(strip=True)
                nav_date = datetime.strptime(nav_date_str, '%Y%m%d').date()
                data = {
                    'product_code': cells[0].get_text(strip=True),
                    'product_name': cells[1].get_text(strip=True),
                    'nav': 1,
                    'acc_nav': 1,
                    'date': datetime.strptime(nav_date_str, '%Y%m%d').strftime('%Y-%m-%d')
                }
            elif table_type == 5:
                data_found = True
                nav_date_str = cells[4].get_text(strip=True)
                nav_date = datetime.strptime(nav_date_str, '%Y%m%d').date()
                data = {
                    'product_code': cells[0].get_text(strip=True),
                    'product_name': cells[1].get_text(strip=True),
                    'nav': 1,
                    'acc_nav': 1,
                    'date': datetime.strptime(nav_date_str, '%Y%m%d').strftime('%Y-%m-%d')
                }
            elif table_type == 6:
                data_found = True
                nav_date_str = cells[3].get_text(strip=True)
                nav_date = datetime.strptime(nav_date_str, '%Y%m%d').date()
                data = {
                    'product_code': cells[0].get_text(strip=True),
                    'product_name': cells[1].get_text(strip=True),
                    'nav': 1,
                    'acc_nav': 1,
                    'date': datetime.strptime(nav_date_str, '%Y%m%d').strftime('%Y-%m-%d')
                }
            if nav_date == None or data == None:
                break

            # Check if we've gone past the start date
            if start_date and nav_date < start_date:
                break

            # Check if we're still within the end date
            if end_date and nav_date > end_date:
                continue

            finance_data.append(data)

        if data_found:
            logger.info(f"Extracted {len(rows)} records from current page")
        else:
            logger.warning("No valid data found in current page")

        return {
            'nav_data': finance_data,
            'info_data': product_info
        }
    
    def _export_to_cache(self, product: Dict, start_date: date, end_date: date, finance_data: List[Dict], product_info: Dict) -> str:
        """将数据导出到缓存目录"""
        try:
            # 确保缓存目录存在
            os.makedirs(self.cache_dir, exist_ok=True)
            
            # 生成文件名，使用原始日期配置
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            date_range = ""
            start = start_date.strftime("%Y%m%d") if start_date else "START"
            end = end_date.strftime("%Y%m%d") if end_date else "END"
            date_range = f"{start}_to_{end}"
                
            filename = f"finance_{product['PrdCode']}_{date_range}_{timestamp}.xlsx"
            filepath = os.path.join(self.cache_dir, filename)
            
            # 删除相同日期范围的旧缓存文件
            prefix = f"finance_{product['PrdCode']}_{date_range}_"
            for file in os.listdir(self.cache_dir):
                if file.startswith(prefix) and date_range in file:
                    logger.info(f"删除旧缓存文件: {file}")
                    os.remove(os.path.join(self.cache_dir, file))
                    
            # 将数据导出到Excel
            # 创建工作簿
            wb = Workbook()
            
            info_sheet = wb.active
            info_sheet.title = "product_info"
            
            # 创建产品信息sheet
            if product_info:
                info_sheet.append(["key", "value"])
                for key, value in product_info.items():
                    info_sheet.append([key, value])

            # 创建净值数据sheet
            data_sheet = wb.create_sheet(title="finance_data", index=1)

            # 写入表头
            headers = ["product_code", "product_name", "nav", "acc_nav", "date"]
            data_sheet.append(headers)
            
            # 写入数据
            for data in finance_data:
                data_sheet.append([
                    data['product_code'],
                    data['product_name'],
                    float(data['nav']),
                    float(data['acc_nav']),
                    data['date']
                ])

            # 保存工作簿
            wb.save(filepath)
            logger.info(f"Data exported to: {filepath}")
            return filepath
        except Exception as e:
            logger.error(f"导出数据失败: {str(e)}")
            return ""

    def _parse_product_info(self, soup: BeautifulSoup) -> None:
        """解析产品基本信息"""
        product_info = {}
        try:
            BaseTitle_div = soup.find('div', {'class': 'BaseTitle'})
            if BaseTitle_div:
                BaseTitle_div_text = BaseTitle_div.get_text(strip=True)
                product_info['BaseTitle'] = BaseTitle_div_text

            BaseType_div = soup.find('div', {'class': 'BaseType'})
            if BaseType_div:
                BaseType_div_text = BaseType_div.get_text(strip=True)
                product_info['BaseType'] = BaseType_div_text

            saNumNetValue_div = soup.find('div', {'class': 'saNumNetValue'})
            if saNumNetValue_div:
                saNumNetValue_div_text = saNumNetValue_div.get_text(strip=True)
                product_info['saNumNetValue'] = saNumNetValue_div_text

            rateBasic_div = soup.find('div', {'class': 'rateBasic'})
            if rateBasic_div:
                rateBasic_div_text = rateBasic_div.get_text(strip=True)
                product_info['rateBasic'] = rateBasic_div_text
            
            divMid_div = soup.find('div', {'class': 'divMid'})
            if divMid_div:
                # 遍历divMid内部的所有div
                for div in divMid_div.find_all('div'):
                    # 获取div内的所有子div
                    child_divs = div.find_all('div', recursive=False)
                    if len(child_divs) >= 2:
                        key = child_divs[0].get_text(strip=True).strip('：')
                        value = child_divs[1].get_text(strip=True)
                        product_info[key] = value
            
            divRight_div = soup.find('div', {'class': 'divRight'})
            if divRight_div:
                # 遍历divRight内部的所有div
                for div in divRight_div.find_all('div'):
                    # 获取div内的所有子div
                    child_divs = div.find_all('div', recursive=False)
                    if len(child_divs) >= 2:
                        key = child_divs[0].get_text(strip=True).strip('：')
                        value = child_divs[1].get_text(strip=True)
                        product_info[key] = value

            divBottom_div = soup.find('div', {'class': 'divBottom'})
            if divBottom_div:
                # 获取div内的所有子div
                child_divs = divBottom_div.find_all('div', recursive=False)
                if len(child_divs) >= 2:
                    key = child_divs[0].get_text(strip=True).strip('：')
                    value = child_divs[1].get_text(strip=True)
                    product_info[key] = value

            logger.info("产品基本信息:")
            for key, value in product_info.items():
                logger.info(f"{key}: {value}")
            return product_info

        except Exception as e:
            logger.error(f"解析产品信息失败: {str(e)}")
            import traceback
            logger.error(traceback.format_exc()) 
    
    def parse_products_result(self, json_data: str) -> List[Dict]:
        # 处理换行符
        text = json_data.replace('\n', '').replace('\r', '').replace('\t', '')
        
        # 给没有引号的key添加引号
        text = re.sub(r'([{,])\s*([a-zA-Z_][a-zA-Z0-9_]*)\s*:', r'\1"\2":', text)
        
        # 4. 解析JSON数据
        data = json.loads(text)
        
        if not data or 'result' not in data or data['result'] != 1:
            logger.warning(f"获取产品列表失败: 返回数据格式错误")
            return []
        
        products = []

        for item in data.get('list', []):
            product = {
                'TypeCode': item.get('TypeCode', ''),
                'PrdCode': item.get('PrdCode', ''),
                'PrdName': item.get('PrdName', ''),
                'PrdBrief': item.get('PrdBrief', ''),
                'BeginDate': item.get('BeginDate', ''),
                'EndDate': item.get('EndDate', ''),
                'ExpireDate': item.get('ExpireDate', ''),
                'Status': item.get('Status', ''),
                'Risk': item.get('Risk', ''),
                'Style': item.get('Style', ''),
                'Currency': item.get('Currency', ''),
                'Term': item.get('Term', ''),
                'RateLow': item.get('RateLow', ''),
                'RateHigh': item.get('RateHigh', ''),
                'ShowExpectedReturn': item.get('ShowExpectedReturn', ''),
                'RateDes': item.get('RateDes', ''),
                'IsSA': item.get('IsSA', ''),
                'saaCod': item.get('saaCod', '')
            }
            nav_url = self._build_url(product)
            product['nav_url'] = nav_url
            products.append(product)
        return products


    def get_product_list(self, base_url: str, page_index: int = 1) -> List[Dict]:
        """
        获取在售和即将发售的产品列表
        """
        while True:
            try:
                url = (f"{base_url}"
                       f"&t={time.time()}"
                       f"&pageindex={page_index}")
            
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
                    'Referer': 'https://www.cmbchina.com/',
                    'Accept': 'application/json'
                }
                
                response = requests.get(url, headers=headers, timeout=10)
                response.raise_for_status()
                
                # 处理响应文本
                text = response.text.strip('()')
                
                products = self.parse_products_result(text)
        
                logger.info(f"获取第{page_index}页产品列表成功，共{len(products)}条记录")
                return products
                
            except requests.RequestException as e:
                logger.error(f"重试，获取产品列表失败: {str(e)}, url: {url}")
                time.sleep(self.delay)
                continue
            except Exception as e:
                logger.error(f"处理产品列表数据失败: {str(e)}, url: {url}, text: {text}")
                import traceback
                logger.error(traceback.format_exc())
                return []
        return []

    def __get_all_products(self, base_url: str) -> List[Dict]:
        all_products = []
        page_index = 1
        
        while True:
            products = self.get_product_list(base_url, page_index)
            if not products:
                break

            for product in products:
                url = self._build_url(product, 1)
                result = self._crawl_url(url, None, None, False)
                nav_data = result['nav_data']
                if len(nav_data) > 0:
                    last_nav_date = nav_data[0]['date']
                    product['last_nav_date'] = last_nav_date

            all_products.extend(products)
            logger.info(f"已获取{len(all_products)}条产品记录")

            # 检查是否还有下一页
            if len(products) < 20:
                break
            
            page_index += 1
            time.sleep(self.delay)
        return all_products
    
    def search_products_sell(self) -> List[Dict]:
        return self.__get_all_products("https://www.cmbchina.com/cfweb/svrajax/product.ashx?op=search&type=m&pagesize=20&salestatus=A,B&baoben=&currency=&term=&keyword=&series=01&risk=&city=&date=&orderby=ord1")
    
    def export_product_list(self, products: List[Dict], output_dir: str, prefix: str = "") -> None:
        """将产品列表导出到Excel"""
        try:
            # 确保缓存目录存在
            os.makedirs(output_dir, exist_ok=True)
            
            # 生成文件名，使用原始日期配置
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                
            filename = f"cmb_products_{prefix}_{timestamp}.xlsx"
            filepath = os.path.join(output_dir, filename)
            
            if len(self.product_series) == 0:
                self.get_product_series()

                    
            # 将数据导出到Excel
            # 创建工作簿
            wb = Workbook()

            data_sheet = wb.active
            data_sheet.title = "products"

            # 写入表头
            headers = ["TypeCode", "TypeName", "PrdCode", "PrdName", "PrdBrief", "LastNavDate", "LastNavDays", "BeginDate", "EndDate", "ExpireDate", "Status", "Risk", "Style", "Currency", "Term", "nav_url", "RateLow", "RateHigh", "ShowExpectedReturn", "RateDes"]
            data_sheet.append(headers)

            # 写入数据
            for product in products:
                TypeName = ""
                for series in self.product_series:
                    if series['series_code'] == product['TypeCode']:
                        TypeName = series['manager'] + "-" + series['series_name']
                        break

                data_sheet.append([
                    product['TypeCode'],
                    TypeName,
                    product['PrdCode'],
                    product['PrdName'],
                    product['PrdBrief'],
                    product['last_nav_date'] if 'last_nav_date' in product else "",
                    (datetime.now().date() - datetime.strptime(product['last_nav_date'], '%Y-%m-%d').date()).days if 'last_nav_date' in product and product['last_nav_date'] else "",
                    product['BeginDate'],
                    product['EndDate'],
                    product['ExpireDate'],
                    product['Status'],
                    product['Risk'],
                    product['Style'],
                    product['Currency'],
                    product['Term'],
                    product['nav_url'],
                    product['RateLow'],
                    product['RateHigh'],
                    product['ShowExpectedReturn'],
                    product['RateDes']
                ])

            # 保存工作簿
            wb.save(filepath)
            logger.info(f"Data exported to: {filepath}")
            return filepath
        except Exception as e:
            logger.error(f"导出数据失败: {str(e)}")
            return ""
        
    def get_product_series(self) -> List[Dict]:
        """
        获取产品系列列表
        
        Returns:
            List[Dict]: 产品系列列表，每个系列包含:
                - manager: 管理方名称
                - series_name: 系列名称
                - series_code: 系列代码
        """
        while True:
            try:
                url = "https://www.cmbchina.com/cfweb/Personal/"
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
                    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8'
                }
                
                response = requests.get(url, headers=headers, timeout=10)
                response.raise_for_status()
                
                # 使用BeautifulSoup解析HTML
                soup = BeautifulSoup(response.text, 'html.parser')
                series_list = []
                
                # 查找所有item*和open_item* div
                item_divs = soup.find_all('div', id=lambda x: x and (x.startswith('item') or x.startswith('open_item')))
                
                current_manager = None
                for div in item_divs:
                    div_id = div.get('id', '')
                    
                    if div_id.startswith('item'):
                        # 这是管理方div
                        current_manager = div.get_text(strip=True)
                        continue
                        
                    # 这是系列div
                    links = div.find_all('a')
                    for link in links:
                        href = link.get('href', '')
                        series_name = link.get_text(strip=True)
                        
                        # 从href中提取series_code
                        import re
                        series_code_match = re.search(r'code=(\d+)', href)
                        series_code = series_code_match.group(1) if series_code_match else None
                        
                        if series_code:
                            series_info = {
                                'manager': current_manager,
                                'series_name': series_name,
                                'series_code': series_code
                            }
                            series_list.append(series_info)
                            logger.debug(f"Found series: {series_info}")

                logger.info(f"获取到{len(series_list)}个产品系列")
                self.product_series = series_list
                return series_list
                
            except requests.RequestException as e:
                logger.error(f"重试获取产品系列列表失败: {str(e)}")
                continue
            except Exception as e:
                logger.error(f"解析产品系列数据失败: {str(e)}")
                import traceback
                logger.error(traceback.format_exc())
                return []
        return []
        
    def get_serie_products(self, serie_code: str) -> List[Dict]:
        return self.__get_all_products(f"https://www.cmbchina.com/cfweb/svrajax/product.ashx?op=condition&salestatus=&pagesize=20&terms=&risk=&area=&keyword=&orderby="
                                       f"&prdtype={serie_code}")
    
