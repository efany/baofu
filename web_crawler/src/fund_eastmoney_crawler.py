import requests
from bs4 import BeautifulSoup
from typing import Set, List, Dict
import pandas as pd
import time
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta
import os
import re
from loguru import logger

class FundEastmoneyCrawler:
    def __init__(self, fund_code: str, start_date: str = None, end_date: str = None, per_page: int = 20, cache_dir: str = "cache/eastmoney"):
        """
        初始化基金数据爬虫
        
        Args:
            fund_code: 基金代码
            start_date: 开始日期，格式：YYYY-MM-DD
            end_date: 结束日期，格式：YYYY-MM-DD
            per_page: 每页记录数
            cache_dir: 缓存目录
        """
        self.fund_code = fund_code
        self.start_date = start_date
        self.end_date = end_date
        self.per_page = per_page
        self.cache_dir = cache_dir
        self.nav_data = []  # 净值数据
        self.shares_data = []  # 份额数据
        self.fee_data = {}  # 费率数据
        self.session = requests.Session()
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        }

    def _make_request(self, url: str, max_retries: int = 3) -> requests.Response:
        """发送请求并处理重试"""
        retry_count = 0
        while retry_count < max_retries:
            try:
                response = requests.get(
                    url, 
                    headers=self.headers,
                    timeout=30,
                    stream=True  # 使用流式传输
                )
                response.raise_for_status()
                
                # 读取完整的响应内容
                content = b""
                for chunk in response.iter_content(chunk_size=8192):
                    if chunk:
                        content += chunk
                
                # 创建新的响应对象，包含完整内容
                new_response = requests.Response()
                new_response._content = content
                new_response.status_code = response.status_code
                new_response.headers = response.headers
                new_response.encoding = response.encoding
                
                return new_response
                
            except (requests.exceptions.RequestException, 
                    requests.exceptions.ChunkedEncodingError) as e:
                retry_count += 1
                wait_time = retry_count * 2
                logger.error(f"请求失败，{wait_time}秒后进行第{retry_count}次重试...")
                logger.error(f"错误信息: {str(e)}")
                time.sleep(wait_time)
                
                if retry_count == max_retries:
                    raise

    def get_nav_data(self) -> List[Dict[str, str]]:
        """获取基金净值数据"""
        nav_data = []
        max_retries = 3
        retry_count = 0
        
        try:
            page = 1
            reached_start = False  # 是否到达开始日期
            
            while True:
                logger.info(f"正在获取第 {page} 页数据...")
                
                try:
                    # 构建URL（不添加日期参数）
                    url = (f'http://fund.eastmoney.com/f10/F10DataApi.aspx?type=lsjz'
                          f'&code={self.fund_code}&page={page}&per={self.per_page}')
                    logger.info(f"url: {url}")
                    response = self._make_request(url)
                    response.encoding = 'utf-8'

                    # 解析数据
                    soup = BeautifulSoup(response.text, 'html.parser')
                    table = soup.find('table', {'class': 'w782 comm lsjz'})

                    if not table:
                        if retry_count < max_retries:
                            retry_count += 1
                            wait_time = retry_count * 2
                            logger.info(f"未找到数据表格，{wait_time}秒后进行第{retry_count}次重试...")
                            time.sleep(wait_time)
                            continue
                        break
                    
                    rows = table.find_all('tr')[1:]  # 跳过表头
                    if not rows:
                        break
                    
                    # 重置重试计数
                    retry_count = 0
                    has_data = False
                    
                    for row in rows:
                        cells = row.find_all('td')
                        if len(cells) >= 7:
                            date = cells[0].get_text(strip=True)
                            has_data = True
                            # 检查日期范围（数据是倒序的
                            if self.end_date and date > self.end_date:
                                continue
                            
                            if self.start_date and date < self.start_date:
                                reached_start = True
                                break  # 已经超过开始日期，不需要继续获取更早的数据
                            
                            # 只有在日期范围内的数据才添加到结果中
                            if (not self.start_date or date >= self.start_date) and \
                               (not self.end_date or date <= self.end_date):
                                nav_data.append({
                                    'date': date,
                                    'nav': cells[1].get_text(strip=True),
                                    'acc_nav': cells[2].get_text(strip=True),
                                    'growth_rate': cells[3].get_text(strip=True).replace('%', ''),
                                    'buy_status': cells[4].get_text(strip=True),
                                    'sell_status': cells[5].get_text(strip=True),
                                    'dividend': cells[6].get_text(strip=True)
                                })
                
                    # 如果已经到达开始日期或没有更多数据，则停止获取
                    if reached_start or not has_data:
                        logger.info(f"已经到达开始日期或没有更多数据，停止获取")
                        break
                
                except Exception as e:
                    if retry_count < max_retries:
                        retry_count += 1
                        wait_time = retry_count * 2
                        logger.error(f"处理数据失败，{wait_time}秒后进行第{retry_count}次重试...")
                        logger.error(f"错误信息: {str(e)}")
                        time.sleep(wait_time)
                        continue
                    else:
                        logger.info(f"达到最大重试次数，跳过当前页")
                        break
                
                page += 1
                time.sleep(0.5)  # 添加延迟避免请求过快
            
            logger.info(f"获取到 {len(nav_data)} 条净值数据")
            return nav_data

        except Exception as e:
            logger.error(f"获取净值数据失败: {str(e)}")
            import traceback
            logger.error(traceback.format_exc())
            return nav_data

    def get_fund_shares(self) -> List[Dict[str, str]]:
        """获取基金份额数据"""
        shares_data = []
        try:
            # 构建URL
            url = f'https://fundf10.eastmoney.com/FundArchivesDatas.aspx?type=gmbd&code={self.fund_code}'
            
            # 发送请求
            response = self._make_request(url)
            response.encoding = 'utf-8'
            
            # 提取数据部分
            content_start = response.text.find('content:"') + 9
            content_end = response.text.find('",summary:')
            if content_start > 8 and content_end > 0:
                content = response.text[content_start:content_end]
                
                # 解析表格数据
                soup = BeautifulSoup(content, 'html.parser')
                table = soup.find('table', {'class': 'w782 comm gmbd'})
                if table:
                    # 跳过表头
                    rows = table.find_all('tr')[1:]  
                    for row in rows:
                        cells = row.find_all('td')
                        if len(cells) >= 6:
                            share_data = {
                                'share_date': cells[0].get_text(strip=True),  # 日期
                                'purchase': cells[1].get_text(strip=True),    # 期间申购（亿份）
                                'redeem': cells[2].get_text(strip=True),      # 期间赎回（亿份）
                                'total_share': cells[3].get_text(strip=True), # 期末总份额（亿份）
                                'total_asset': cells[4].get_text(strip=True), # 期末净资产（亿元）
                                'change_rate': cells[5].get_text(strip=True)  # 净资产变动率
                            }
                            shares_data.append(share_data)
                            
                    return shares_data

        except Exception as e:
            logger.error(f"获取基金份额数据失败: {str(e)}")
            import traceback
            logger.error(traceback.format_exc())
            return []
    
    def get_fee_data(self) -> Dict[str, str]:
        """获取基金费率和交易状态数据"""
        try:
            # 构建URL
            url = f'https://fund.eastmoney.com/{self.fund_code}.html'
            response = self._make_request(url)
            response.encoding = 'utf-8'
            
            # 解析HTML
            soup = BeautifulSoup(response.text, 'html.parser')
            
            # 获取费率和交易状态信息
            fee_data = {
                'purchase_rate': '',    # 购买费率
                'actual_rate': '',      # 实际费率
                'discount': '',         # 折扣
                'manage_rate': '',      # 管理费率
                'custody_rate': '',     # 托管费率
                'sale_rate': '',        # 销售服务费率
                'purchase_status': '',  # 申购状态
                'redeem_status': ''     # 赎回状态
            }
            
            # 获取申购费率信息
            rate_text = soup.find('span', text=re.compile(r'购买手续费：'))
            if rate_text and rate_text.parent:
                rate_info = rate_text.parent.get_text()
                
                # 解析费率信息
                rate_matches = re.findall(r'([\d.]+)%', rate_info)
                if rate_matches:
                    fee_data['purchase_rate'] = rate_matches[0]  # 第一个数字是原费率
                    if len(rate_matches) > 1:
                        fee_data['actual_rate'] = rate_matches[1]  # 第二个数字是实际费率
                    else:
                        fee_data['actual_rate'] = rate_matches[0]  # 只有一个数字时，原费率即实际费率
                
                # 提取折扣信息
                discount_match = re.search(r'(\d+)折', rate_info)
                if discount_match:
                    fee_data['discount'] = discount_match.group(1)
                    # 如果没有明确的实际费率，根据折扣计算
                    if not fee_data['actual_rate'] and fee_data['purchase_rate']:
                        try:
                            discount = float(discount_match.group(1)) / 10
                            original_rate = float(fee_data['purchase_rate'])
                            fee_data['actual_rate'] = str(round(original_rate * discount, 4))
                        except (ValueError, TypeError):
                            pass
            
            # 获取交易状态
            trade_status = soup.find('div', {'class': 'fundDetail-main'})
            if trade_status:
                # 查找交易状态文本
                status_text = trade_status.get_text()
                
                # 提取申购状态
                purchase_match = re.search(r'(限大额|暂停申购|开放申购|场内买入).*?(?=\s|开放赎回|暂停赎回|场内卖出|$)', status_text)
                if purchase_match:
                    fee_data['purchase_status'] = purchase_match.group(0).strip()
                    # 如果是限大额，提取限额信息
                    # 限大额 （单日累计购买上限3000元） 开放赎回
                    # 限大额 （单日累计购买上限1.00万元） 开放赎回
                    # 如果是限大额，提取限额信息
                    if '限大额' in fee_data['purchase_status']:
                        limit_match = re.search(r'单日累计购买上限([\d.]+)(万元|元)', status_text)
                        if limit_match:
                            limit_amount = limit_match.group(1)
                            unit = limit_match.group(2)
                            if unit == '万元':
                                limit_amount = float(limit_amount) * 10000
                            fee_data['purchase_status'] += f" （单日累计购买上限{int(float(limit_amount))}元）"
                
                # 提取赎回状态
                redeem_match = re.search(r'(开放赎回|暂停赎回|场内卖出)', status_text)
                if redeem_match:
                    fee_data['redeem_status'] = redeem_match.group(0).strip()
            
            # 获取其他费率信息（从基金概况页面）
            fee_url = f'http://fundf10.eastmoney.com/jbgk_{self.fund_code}.html'
            fee_response = self._make_request(fee_url)
            fee_response.encoding = 'utf-8'
            fee_soup = BeautifulSoup(fee_response.text, 'html.parser')
            
            # 查找费率表格
            fee_table = fee_soup.find('table', {'class': 'info w790'})
            if fee_table:
                rows = fee_table.find_all('tr')
                for row in rows:
                    cells = row.find_all('td')
                    for i in range(0, len(cells)-1, 2):
                        label = cells[i].get_text(strip=True)
                        value = cells[i+1].get_text(strip=True)
                        
                        # 处理不同的费率标签
                        if any(x in label for x in ['管理费率', '管理费']):
                            fee_data['manage_rate'] = value.replace('%', '')
                        elif any(x in label for x in ['托管费率', '托管费']):
                            fee_data['custody_rate'] = value.replace('%', '')
                        elif any(x in label for x in ['销售服务费率', '销售服务费']):
                            fee_data['sale_rate'] = value.replace('%', '')
            
            # 确保所有费率都是有效的数字字符串
            for key in ['purchase_rate', 'actual_rate', 'manage_rate', 'custody_rate', 'sale_rate']:
                if not fee_data[key] or fee_data[key] == '--':
                    fee_data[key] = '0'
                # 移除可能的额外字符
                fee_data[key] = re.sub(r'[^0-9.]', '', fee_data[key])
            
            return fee_data
            
        except Exception as e:
            logger.error(f"获取费率数据失败: {str(e)}")
            import traceback
            logger.error(traceback.format_exc())
            return {
                'purchase_rate': '0',
                'actual_rate': '0',
                'discount': '0',
                'manage_rate': '0',
                'custody_rate': '0',
                'sale_rate': '0',
                'purchase_status': '未知',
                'redeem_status': '未知'
            }

    def _find_cached_file(self, cache_dir: str, data_type: str) -> str:
        """查找符合日期范围的缓存文件"""
        try:
            if not os.path.exists(cache_dir):
                return ""
            
            # 查找所有该基金的Excel文件
            prefix = f"fund_{self.fund_code}_"
            files = [f for f in os.listdir(cache_dir) if f.startswith(prefix) and f.endswith('.xlsx')]
            
            # 获取当前配置的日期范围
            current_range = ""
            if self.start_date is not None or self.end_date is not None:
                start = self.start_date if self.start_date is not None else "START"
                end = self.end_date if self.end_date is not None else "END"
                current_range = f"_{start}_to_{end}"
            
            for file in files:
                # 解析文件名中的日期范围
                file_range = file[len(prefix):-20]  # 去掉前缀和时间戳部分
                
                # 直接比较日期范围字符串
                if file_range == current_range:
                    return os.path.join(cache_dir, file)
            
            return ""
            
        except Exception as e:
            logger.error(f"查找缓存文件失败: {str(e)}")
            return ""

    def _load_cached_data(self, cache_dir: str = "output") -> bool:
        """从缓存文件加载数据"""
        wb = None
        try:
            # 获取当前配置的日期范围
            current_range = ""
            if self.start_date is not None or self.end_date is not None:
                start = self.start_date if len(self.start_date) > 0 else "START"
                end = self.end_date if len(self.end_date) > 0 else "END"
                current_range = f"{start}_to_{end}"
            
            # 查找所有该基金的Excel文件
            prefix = f"fund_{self.fund_code}_{current_range}"
            if not os.path.exists(cache_dir):
                return False, None
            
            files = [f for f in os.listdir(cache_dir) 
                    if f.startswith(prefix) and f.endswith('.xlsx')]
            cached_file = files[0] if len(files) > 0 else None
            
            if not cached_file:
                return False, None
            
            logger.info(f"找到缓存文件: {cached_file}")
        
            # 加载缓存文件
            filepath = os.path.join(cache_dir, cached_file)
            wb = load_workbook(filepath, read_only=True)
            
            # 加载净值数据
            if 'nav_data' in wb.sheetnames:
                ws_nav = wb['nav_data']
                headers = [cell.value for cell in ws_nav[1]]
                
                self.nav_data = []
                for row in ws_nav.iter_rows(min_row=2):
                    data = {}
                    for idx, header in enumerate(headers):
                        value = row[idx].value
                        if value is None:
                            value = ''
                        data[header] = str(value)
                    self.nav_data.append(data)
            
            # 加载份额数据
            if 'shares_data' in wb.sheetnames:
                ws_shares = wb['shares_data']
                headers = [cell.value for cell in ws_shares[1]]
                
                self.shares_data = []
                for row in ws_shares.iter_rows(min_row=2):
                    data = {}
                    for idx, header in enumerate(headers):
                        value = row[idx].value
                        if value is None:
                            value = ''
                        data[header] = str(value)
                    self.shares_data.append(data)
            
            # 加载费率数据
            if 'fee_data' in wb.sheetnames:
                ws_fee = wb['fee_data']
                
                self.fee_data = {}
                for row in ws_fee.iter_rows():
                    key = row[0].value
                    value = row[1].value
                    if key is not None:  # 只处理key不为空的行
                        self.fee_data[str(key)] = str(value) if value is not None else ''
            
            # 如果没有设置结束日期，检查是否需要更新数据
            if not self.end_date and self.nav_data:
                latest_date = self.nav_data[0]['date']  # 缓存中最新的数据日期
                logger.info(f"缓存中最新数据日期: {latest_date}")
                return True, latest_date

            return True, None
            
        except Exception as e:
            logger.error(f"加载缓存数据失败: {str(e)}")
            return False, None
        
        finally:
            if wb:
                wb.close()

    def fetch_fund_data(self) -> Dict[str, List[Dict[str, str]]]:
        """
        获取所有基金数据
        
        Returns:
            Dict: 包含净值数据、份额数据和费率数据的字典
            {
                'nav_data': [...],     # 净值数据列表
                'shares_data': [...],  # 份额数据列表
                'fee_data': {...}      # 费率数据字典
            }
        """
        try:
            # 尝试从缓存加载数据
            cache_loaded, latest_date = self._load_cached_data(self.cache_dir)
            
            if cache_loaded:
                logger.info("使用缓存数据")
                
                # 如果没有设置结束日期，检查是否新数据
                if not self.end_date and latest_date:
                    logger.info(f"检查 {latest_date} 之后的新数据...")

                    # 获取当前日期和时间
                    now = datetime.now()
                    today = now.strftime('%Y-%m-%d')
                    
                    # 判断今天是否是工作日（简单判断是否是周末）
                    today_weekday = now.weekday()
                    is_weekend = today_weekday >= 5
                    
                    # 如果是周末，最后一个完整交易日应该是上周五
                    if is_weekend:
                        days_to_subtract = today_weekday - 4  # 4 represents Friday
                        last_trading_day = (now - timedelta(days=days_to_subtract)).strftime('%Y-%m-%d')
                    else:
                        # 如果是工作日，根据当前时间判断最后一个完整交易日
                        current_hour = now.hour
                        if current_hour >= 16:  # 如果是下午4点后
                            last_trading_day = today
                        else:
                            last_trading_day = (now - timedelta(days=1)).strftime('%Y-%m-%d')
                    
                    # 如果缓存的最新日期不是最后一个交易日，说明需要更新数据
                    if latest_date < last_trading_day:
                        logger.info(f"缓存数据需要更新: 最新数据日期 {latest_date}, 最后交易日 {last_trading_day}")
                    else:
                        logger.info(f"缓存数据已是最新: {latest_date}")
                        return {
                            'nav_data': self.nav_data,
                            'shares_data': self.shares_data,
                            'fee_data': self.fee_data
                        }
                    
                    # 临时设置开始日期为缓存中最新日期的下一天
                    original_start_date = self.start_date
                    self.start_date = latest_date
                    
                    # 获取新数据
                    new_nav_data = self.get_nav_data()
                    if new_nav_data:
                        # 移除重复的日期数据
                        new_nav_data = [data for data in new_nav_data 
                                      if data['date'] > latest_date]
                        if new_nav_data:
                            logger.info(f"获取到 {len(new_nav_data)} 条新数据")
                            # 将新数据添加到缓存数据前面
                            self.nav_data = new_nav_data + self.nav_data

                            # 更新份额数据
                            self.shares_data = self.get_fund_shares()
                            logger.info(f"获取到 {len(self.shares_data)} 条份额数据")
                        else:
                            logger.info("没有新数据")
                    
                    # 恢复原始开始日期
                    self.start_date = original_start_date

                        # 获取份额数据
                    self.shares_data = self.get_fund_shares()
                    logger.info(f"共获取 {len(self.shares_data)} 条份额数据")
                
                    # 获取费率数据
                    logger.info("获取费率数据...")
                    self.fee_data = self.get_fee_data()
                    if self.fee_data:
                        logger.info("费率数据获取成功")
                    else:
                        logger.info("费率数据获取失败")
            else:
                # 如果没有缓存或缓存不满足要求，重新获取数据
                logger.info("重新获取数据...")
                
                # 获取净值数据
                self.nav_data = self.get_nav_data()
                logger.info(f"共获取 {len(self.nav_data)} 条净值数据")
                
                # 获取份额数据
                self.shares_data = self.get_fund_shares()
                logger.info(f"共获取 {len(self.shares_data)} 条份额数据")
            
                # 获取费率数据
                logger.info("获取费率数据...")
                self.fee_data = self.get_fee_data()
                if self.fee_data:
                    logger.info("费率数据获取成功")
                else:
                    logger.info("费率数据获取失败")
            
            self.export_to_cache()

            return {
                'nav_data': self.nav_data,
                'shares_data': self.shares_data,
                'fee_data': self.fee_data
            }
            
        except Exception as e:
            logger.error(f"获取数据失败: {str(e)}")
            import traceback
            logger.error(traceback.format_exc())
            return {
                'nav_data': [],
                'shares_data': [],
                'fee_data': {}
            }
        
    def export_to_cache(self) -> str:
        """
        将数据导出到缓存目录
        
        Returns:
            str: 缓存文件路径
        """
        try:
            # 确保缓存目录存在
            os.makedirs(self.cache_dir, exist_ok=True)
            
            # 生成文件名，使用原始日期配置
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            date_range = ""
            if self.start_date is not None or self.end_date is not None:
                start = self.start_date if len(self.start_date) > 0 else "START"
                end = self.end_date if len(self.end_date) > 0 else "END"
                date_range = f"{start}_to_{end}"
            
            filename = f"fund_{self.fund_code}_{date_range}_{timestamp}.xlsx"
            filepath = os.path.join(self.cache_dir, filename)

            # 删除相同日期范围的旧缓存文件
            prefix = f"fund_{self.fund_code}_"
            for file in os.listdir(self.cache_dir):
                if file.startswith(prefix) and file.endswith('.xlsx'):
                    # 解析文件名中的日期范围
                    file_range = file[len(prefix):-21]  # 去掉前缀和时间戳部分
                    logger.info(f"缓存文件: {file}, 日期范围: {file_range}")
                    # 如果日期范围相同，删除旧文件
                    if file_range == date_range:
                        old_file = os.path.join(self.cache_dir, file)
                        try:
                            os.remove(old_file)
                            logger.info(f"删除旧缓存文件: {file}")
                        except Exception as e:
                            logger.error(f"删除旧缓存文件失败: {str(e)}")
            
            # 创建工作簿
            wb = Workbook()
            
            # 创建并保存新文件
            if self.nav_data:
                # 使用第一个工作表保存净值数据
                ws_nav = wb.active
                ws_nav.title = "nav_data"
                nav_df = pd.DataFrame(self.nav_data)
                # 写入表头
                for c_idx, column in enumerate(nav_df.columns, 1):
                    ws_nav.cell(row=1, column=c_idx, value=column)
                # 从第2行开始写入数据
                for r_idx, row in enumerate(nav_df.values, 2):
                    for c_idx, value in enumerate(row, 1):
                        ws_nav.cell(row=r_idx, column=c_idx, value=value)
            
            if self.shares_data:
                # 创建新工作表保存份额数据
                ws_shares = wb.create_sheet("shares_data")
                shares_df = pd.DataFrame(self.shares_data)
                # 写入表头
                for c_idx, column in enumerate(shares_df.columns, 1):
                    ws_shares.cell(row=1, column=c_idx, value=column)
                for r_idx, row in enumerate(shares_df.values, 2):
                    for c_idx, value in enumerate(row, 1):
                        ws_shares.cell(row=r_idx, column=c_idx, value=value)
            
            if self.fee_data:
                # 创建新工作表保存费率数据
                ws_fee = wb.create_sheet("fee_data")
                
                # 写入费率数据
                for row, (key, value) in enumerate(self.fee_data.items(), 1):
                    ws_fee.cell(row=row, column=1, value=key)
                    ws_fee.cell(row=row, column=2, value=value)
            
            # 如果没有任何数据，至少创建一个空工作表
            if not (self.nav_data or self.shares_data or self.fee_data):
                ws = wb.active
                ws.title = "empty"
            
            # 保存工作簿
            wb.save(filepath)
            logger.info(f"数据已缓存到: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"导出缓存失败: {str(e)}")
            import traceback
            logger.error(traceback.format_exc())
            return ""

    def export_to_excel(self, output_dir: str = "output/eastmoney") -> str:
        """导出数据到Excel"""
        if not self.nav_data and not self.shares_data:
            logger.warning("没有数据可导出")
            return ""

        try:
            os.makedirs(output_dir, exist_ok=True)
            
            # 生成文件名，使用原始日期配置
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            date_range = ""
            if self.start_date is not None or self.end_date is not None:
                start = self.start_date if len(self.start_date) > 0 else "START"
                end = self.end_date if len(self.end_date) > 0 else "END"
                date_range = f"_{start}_to_{end}"
            
            filename = f"fund_{self.fund_code}{date_range}_{timestamp}.xlsx"
            filepath = os.path.join(output_dir, filename)
            
            # 创建并保存新文件
            wb = Workbook()
            
            # 创建净值数据表
            if self.nav_data:
                ws_nav = wb.active
                ws_nav.title = "净值数据"
                
                # 写入净值数据表头
                nav_headers = ["日期", "单位净值", "累计净值", "日增长率(%)", 
                             "申购状态", "赎回状态", "分红送配"]
                for col, header in enumerate(nav_headers, 1):
                    ws_nav.cell(row=1, column=col, value=header)
                
                # 写入净值数据
                for row, data in enumerate(self.nav_data, 2):
                    ws_nav.cell(row=row, column=1, value=data['date'])
                    ws_nav.cell(row=row, column=2, value=data['nav'])
                    ws_nav.cell(row=row, column=3, value=data['acc_nav'])
                    ws_nav.cell(row=row, column=4, value=data['growth_rate'] if data['growth_rate'] else None)
                    ws_nav.cell(row=row, column=5, value=data['buy_status'])
                    ws_nav.cell(row=row, column=6, value=data['sell_status'])
                    ws_nav.cell(row=row, column=7, value=data['dividend'])
            
            # 创建份额数据表
            if self.shares_data:
                ws_shares = wb.create_sheet("份额数据")
                
                # 写入份额数据表头
                shares_headers = [
                    "统计日期", 
                    "期间申购(亿份)", 
                    "期间赎回(亿份)", 
                    "期末总份额(亿份)", 
                    "期末净资产(亿元)", 
                    "净资产变动率(%)"
                ]
                for col, header in enumerate(shares_headers, 1):
                    ws_shares.cell(row=1, column=col, value=header)
                
                def safe_float(value: str) -> float:
                    """安全地转换字符串到浮点数"""
                    try:
                        if value == '---' or not value:
                            return None
                        return float(value)
                    except ValueError:
                        return None
                
                # 写入份额数据
                for row, data in enumerate(self.shares_data, 2):
                    ws_shares.cell(row=row, column=1, value=data['share_date'])
                    ws_shares.cell(row=row, column=2, value=safe_float(data['purchase']))
                    ws_shares.cell(row=row, column=3, value=safe_float(data['redeem']))
                    ws_shares.cell(row=row, column=4, value=safe_float(data['total_share']))
                    ws_shares.cell(row=row, column=5, value=safe_float(data['total_asset']))
                    # 处理百分比值
                    change_rate = data['change_rate'].replace('%', '')
                    ws_shares.cell(row=row, column=6, value=safe_float(change_rate))

            # 创建费率数据表
            ws_fee = wb.create_sheet("费率信息")
            
            # 写入费率数据
            fee_items = [
                ("原始申购费率", 'purchase_rate'),
                ("实际申购费率", 'actual_rate'),
                ("费率折扣", 'discount'),
                ("管理费率", 'manage_rate'),
                ("托管费率", 'custody_rate'),
                ("销售服务费率", 'sale_rate'),
                ("申购状态", 'purchase_status'),
                ("赎回状态", 'redeem_status')
            ]
            
            for row, (label, key) in enumerate(fee_items, 1):
                ws_fee.cell(row=row, column=1, value=label)
                value = self.fee_data.get(key, '--')
                if value and value != '--':
                    if key == 'discount':
                        ws_fee.cell(row=row, column=2, value=f"{value}折")
                    elif key == 'purchase_status' or key == 'redeem_status':
                        ws_fee.cell(row=row, column=2, value=f"{value}")
                    else:
                        ws_fee.cell(row=row, column=2, value=f"{value}%")
                else:
                    ws_fee.cell(row=row, column=2, value=value)
            
            # 调整所有工作表的列宽
            for ws in wb.worksheets:
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column_letter].width = adjusted_width

            # 最后保存新文件
            wb.save(filepath)
            logger.info(f"数据已导出到: {filepath}")
            return filepath

        except Exception as e:
            logger.error(f"导出数据失败: {str(e)}")
            import traceback
            logger.error(traceback.format_exc())
            return ""
  